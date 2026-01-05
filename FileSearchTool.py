# -*- coding: utf-8 -*- # 指定編碼
# -------------------------------------------------
# FileSearchApp v1.6.0 (Added Keyword OR Logic) # MODIFIED version and description
# Description: Deep search content within Excel and PDF files with AND/OR keyword logic.
# Author: Wesley Chang 
# Date: 2025-May (Corrected Year)
# -------------------------------------------------

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import threading
import queue
import re
import time
import platform
import warnings
# 新增導入 sys 模組
import sys
# 新增導入 subprocess 模組
import subprocess

# Suppress openpyxl data validation warnings
warnings.filterwarnings("ignore", category=UserWarning,
                       message="Data Validation extension is not supported and will be removed")

# --- 文件讀取模組 ---
try:
    import openpyxl
    # 為了 .xls 的列名轉換也導入 utils
    from openpyxl.utils import get_column_letter as get_excel_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    # 如果 openpyxl 都沒有，get_excel_column_letter 也不可用
    def get_excel_column_letter(idx): # 提供一個備用函數， although 效果不好
        if idx < 1: return '#'
        s = ''
        while idx > 0:
            idx, r = divmod(idx - 1, 26)
            s = chr(65 + r) + s
        return s
    print("警告：未安裝 openpyxl，無法搜尋 .xlsx 檔案。請執行 'pip install openpyxl'")

try:
    import xlrd
    # 嘗試導入 xlrd.formula 以便使用 cellname，如果失敗則不影響核心功能
    try:
        import xlrd.formula
        HAS_XLRD_FORMULA = True
    except ImportError:
        HAS_XLRD_FORMULA = False
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    HAS_XLRD_FORMULA = False
    print("警告：未安裝 xlrd，無法搜尋 .xls 檔案。請執行 'pip install xlrd'")

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False
    print("警告：未安裝 PyMuPDF (fitz)，無法搜尋 .pdf 檔案。請執行 'pip install PyMuPDF'")

# --- 圖像處理模組 (Pillow) ---
try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("提示：未安裝 Pillow，無法顯示作者頭像。可選安裝：'pip install Pillow'")
# --- --- --- ---

# 新增 resource_path 輔助函數
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        # In development mode, _MEIPASS is not set
        # 使用 __file__ 來獲取當前腳本的目錄
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

class FileSearchApp:
    def __init__(self, master):
        self.master = master
        # MODIFIED version number
        master.title("檔案內容深度搜尋工具 v1.6.0 - 尋找內容含有關鍵字的檔案")
        master.geometry("850x650") # Increased height slightly for new radio buttons and author info

        try:
            icon_path = resource_path("icon.ico")
            master.iconbitmap(icon_path)
        except Exception as e:
            print(f"提示：無法設定應用程式圖示: {e}")

        self.style = ttk.Style()
        available_themes = self.style.theme_names()
        if platform.system() == "Windows":
            if 'vista' in available_themes: self.style.theme_use('vista')
            elif 'clam' in available_themes: self.style.theme_use('clam')
            else: self.style.theme_use('default')
        else: self.style.theme_use('default')

        self.search_thread = None
        self.stop_search_flag = threading.Event()
        self.update_queue = queue.Queue()
        self.result_details = {}
        self.author_image_tk = None

        self.folder_path = tk.StringVar()
        self.keyword1 = tk.StringVar()
        self.keyword2 = tk.StringVar()
        self.keyword_logic = tk.StringVar(value="AND") # MODIFIED: Added for AND/OR logic, default AND
        self.whole_word = tk.BooleanVar()
        self.case_sensitive = tk.BooleanVar()
        self.content_only = tk.BooleanVar(value=True)
        self.file_type_filter = tk.StringVar(value="Both")
        self.include_subfolders = tk.BooleanVar(value=True)

        self.supported_extensions = {
            'Excel': ['.xlsx', '.xls'],
            'PDF': ['.pdf'],
            'Both': ['.xlsx', '.xls', '.pdf']
        }

        self.create_widgets()
        self.master.after(100, self.process_queue)

    def create_widgets(self):
        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))

        folder_frame = ttk.LabelFrame(top_frame, text=" 1. 選擇資料夾 ", padding="5")
        folder_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        self.folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_path, width=70)
        self.folder_entry.pack(side=tk.LEFT, padx=(0, 5), fill=tk.X, expand=True)
        self.browse_button = ttk.Button(folder_frame, text="瀏覽...", command=self.browse_folder, width=10)
        self.browse_button.pack(side=tk.LEFT)

        keyword_options_frame = ttk.Frame(top_frame)
        keyword_options_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        keyword_frame = ttk.LabelFrame(keyword_options_frame, text=" 2. 輸入關鍵字 ", padding="5")
        keyword_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10), anchor=tk.N)
        kw1_label = ttk.Label(keyword_frame, text="關鍵字 1:")
        kw1_label.grid(row=0, column=0, padx=5, pady=3, sticky=tk.W)
        self.keyword1_entry = ttk.Entry(keyword_frame, textvariable=self.keyword1, width=30)
        self.keyword1_entry.grid(row=0, column=1, padx=5, pady=3, sticky=tk.W)

        # MODIFIED: Keyword 2 label
        kw2_label = ttk.Label(keyword_frame, text="關鍵字 2 (可選):")
        kw2_label.grid(row=1, column=0, padx=5, pady=3, sticky=tk.W)
        self.keyword2_entry = ttk.Entry(keyword_frame, textvariable=self.keyword2, width=30)
        self.keyword2_entry.grid(row=1, column=1, padx=5, pady=3, sticky=tk.W)

        # --- START: MODIFIED - Keyword Logic Radiobuttons ---
        logic_frame = ttk.Frame(keyword_frame)
        logic_frame.grid(row=2, column=0, columnspan=2, pady=(2, 5), sticky=tk.W, padx=5)

        self.rb_logic_and = ttk.Radiobutton(logic_frame, text="AND (同時包含)", variable=self.keyword_logic, value="AND")
        self.rb_logic_and.pack(side=tk.LEFT, padx=(0,5))

        self.rb_logic_or = ttk.Radiobutton(logic_frame, text="OR (包含任一)", variable=self.keyword_logic, value="OR")
        self.rb_logic_or.pack(side=tk.LEFT, padx=5)
        # --- END: MODIFIED - Keyword Logic Radiobuttons ---


        image_path = resource_path("author_avatar.png")
        author_image_label = None

        keyword_frame.columnconfigure(0, weight=1)
        keyword_frame.columnconfigure(1, weight=1)

        if HAS_PIL:
            try:
                original_image = Image.open(image_path)
                resized_image = original_image.resize((80, 107), Image.Resampling.LANCZOS)
                self.author_image_tk = ImageTk.PhotoImage(resized_image)
                author_image_label = ttk.Label(keyword_frame, image=self.author_image_tk)
            except FileNotFoundError:
                print(f"提示：找不到作者頭像檔案 '{image_path}'。")
            except Exception as e:
                print(f"提示：載入作者頭像時發生錯誤: {e}")

        author_text_label = ttk.Label(keyword_frame, text="Author: Wesley Chang, May 2025", font=("Arial", 10))

        # MODIFIED: author_info_row adjusted due to new logic_frame
        author_info_row = 3

        if author_image_label:
            author_image_label.grid(row=author_info_row, column=0, padx=5, pady=(10, 5))
            author_text_label.grid(row=author_info_row, column=1, padx=5, pady=(10, 5))
        else:
            author_text_label.grid(row=author_info_row, column=0, columnspan=2, padx=5, pady=(10, 5), sticky=tk.N)


        options_frame = ttk.LabelFrame(keyword_options_frame, text=" 3. 搜尋選項 ", padding="5")
        options_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, anchor=tk.N)
        self.whole_word_check = ttk.Checkbutton(options_frame, text="全詞匹配", variable=self.whole_word)
        self.whole_word_check.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        self.case_sensitive_check = ttk.Checkbutton(options_frame, text="區分大小寫", variable=self.case_sensitive)
        self.case_sensitive_check.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.content_only_check = ttk.Checkbutton(options_frame, text="僅搜尋內容*(忽略公式/註解/被隱藏的頁面)", variable=self.content_only)
        self.content_only_check.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
        self.subfolder_check = ttk.Checkbutton(options_frame, text="包含子資料夾", variable=self.include_subfolders)
        self.subfolder_check.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        file_type_label = ttk.Label(options_frame, text="檔案類型:")
        file_type_label.grid(row=2, column=0, padx=10, pady=(10, 0), sticky=tk.W)
        self.rb_both = ttk.Radiobutton(options_frame, text="Excel與PDF格式", variable=self.file_type_filter, value="Both")
        self.rb_both.grid(row=3, column=0, padx=15, pady=2, sticky=tk.W)
        self.rb_excel = ttk.Radiobutton(options_frame, text="僅 Excel (.xlsx, .xls)", variable=self.file_type_filter, value="Excel")
        self.rb_excel.grid(row=4, column=0, padx=15, pady=2, sticky=tk.W)
        self.rb_pdf = ttk.Radiobutton(options_frame, text="僅 PDF (.pdf)", variable=self.file_type_filter, value="PDF")
        self.rb_pdf.grid(row=5, column=0, padx=15, pady=2, sticky=tk.W)

        search_button_frame = ttk.Frame(main_frame)
        search_button_frame.pack(pady=10)
        self.search_button = ttk.Button(search_button_frame, text="開始搜尋", command=self.start_search, width=15)
        self.search_button.pack(side=tk.LEFT, padx=5)
        self.stop_button = ttk.Button(search_button_frame, text="停止搜尋", command=self.stop_search, state=tk.DISABLED, width=15)
        self.stop_button.pack(side=tk.LEFT, padx=5)

        result_frame = ttk.LabelFrame(main_frame, text=" 4. 搜尋結果 (單擊開啟檔案) ", padding="5")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.tree = ttk.Treeview(result_frame, columns=("filepath", "matches", "filetype"), show="headings")
        self.tree.heading("filepath", text="檔案完整路徑")
        self.tree.heading("matches", text="匹配數")
        self.tree.heading("filetype", text="檔案類型")
        self.tree.column("filepath", width=600, anchor=tk.W)
        self.tree.column("matches", width=70, anchor=tk.CENTER)
        self.tree.column("filetype", width=80, anchor=tk.CENTER)

        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        result_frame.grid_rowconfigure(0, weight=1)
        result_frame.grid_columnconfigure(0, weight=1)
        self.tree.bind("<Button-1>", self.on_tree_click)

        status_progress_frame = ttk.Frame(main_frame)
        status_progress_frame.pack(fill=tk.X, pady=(5, 0))
        self.progress_label = ttk.Label(status_progress_frame, text="進度: 0%", width=15, anchor=tk.W)
        self.progress_label.pack(side=tk.LEFT, padx=(5, 5))
        self.progress_bar = ttk.Progressbar(status_progress_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.status_label = ttk.Label(self.master, text="狀態：待命中...", anchor=tk.W, relief=tk.SUNKEN, padding=(5, 2))
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X, padx=0, pady=0)

    def browse_folder(self):
        foldername = filedialog.askdirectory()
        if foldername:
            self.folder_path.set(foldername)
            self.update_status(f"已選擇資料夾: {foldername}")

    def update_status(self, message):
        self.status_label.config(text=f"狀態：{message}")
        self.master.update_idletasks()

    def update_progress(self, value, total_files):
        if total_files > 0:
            percentage = int((value / total_files) * 100)
            self.progress_bar['value'] = percentage
            self.progress_label.config(text=f"進度: {percentage}% ({value}/{total_files})")
        else:
            self.progress_bar['value'] = 0
            self.progress_label.config(text="進度: 0%")
        self.master.update_idletasks()

    def start_search(self):
        folder = self.folder_path.get()
        kw1 = self.keyword1.get().strip()
        if not folder: messagebox.showerror("輸入錯誤", "請先選擇要搜尋的資料夾！"); return
        if not kw1: messagebox.showerror("輸入錯誤", "請至少輸入第一個關鍵字！"); return
        if not os.path.isdir(folder): messagebox.showerror("路徑錯誤", f"資料夾不存在或無效:\n{folder}"); return

        self.stop_search_flag.clear()
        self.tree.delete(*self.tree.get_children())
        self.result_details.clear()
        self.update_status("正在準備搜尋...")
        self.progress_bar['value'] = 0
        self.progress_label.config(text="進度: 0%")
        self.disable_controls()

        self.search_thread = threading.Thread(target=self._search_worker, daemon=True)
        self.search_thread.start()

    def stop_search(self):
        if self.search_thread and self.search_thread.is_alive():
            self.stop_search_flag.set()
            self.update_status("正在停止搜尋...")
            self.stop_button.config(state=tk.DISABLED)

    def disable_controls(self):
        self.search_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.keyword1_entry.config(state=tk.DISABLED)
        self.keyword2_entry.config(state=tk.DISABLED)
        self.rb_logic_and.config(state=tk.DISABLED) # MODIFIED: Disable logic radio
        self.rb_logic_or.config(state=tk.DISABLED)  # MODIFIED: Disable logic radio
        self.whole_word_check.config(state=tk.DISABLED)
        self.case_sensitive_check.config(state=tk.DISABLED)
        self.content_only_check.config(state=tk.DISABLED)
        self.browse_button.config(state=tk.DISABLED)
        self.rb_both.config(state=tk.DISABLED)
        self.rb_excel.config(state=tk.DISABLED)
        self.rb_pdf.config(state=tk.DISABLED)
        self.subfolder_check.config(state=tk.DISABLED)

    def enable_controls(self):
        self.search_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.keyword1_entry.config(state=tk.NORMAL)
        self.keyword2_entry.config(state=tk.NORMAL)
        self.rb_logic_and.config(state=tk.NORMAL) # MODIFIED: Enable logic radio
        self.rb_logic_or.config(state=tk.NORMAL)  # MODIFIED: Enable logic radio
        self.whole_word_check.config(state=tk.NORMAL)
        self.case_sensitive_check.config(state=tk.NORMAL)
        self.content_only_check.config(state=tk.NORMAL)
        self.browse_button.config(state=tk.NORMAL)
        self.rb_both.config(state=tk.NORMAL)
        self.rb_excel.config(state=tk.NORMAL)
        self.rb_pdf.config(state=tk.NORMAL)
        self.subfolder_check.config(state=tk.NORMAL)

    def process_queue(self):
        try:
            while True:
                message = self.update_queue.get_nowait()
                msg_type = message.get("type")
                data = message.get("data")

                if msg_type == "status": self.update_status(data)
                elif msg_type == "progress": self.update_progress(*data)
                elif msg_type == "add_result":
                    filepath, location_info, kw1_count, kw2_count, file_type = data
                    total_matches = kw1_count + kw2_count
                    item_id = self.tree.insert("", tk.END, values=(filepath, total_matches, file_type))
                    self.result_details[item_id] = (filepath, location_info, kw1_count, kw2_count)
                    self.sort_results_by_matches()
                elif msg_type == "search_complete":
                    total_found = data
                    self.update_status(f"搜尋完成，找到 {total_found} 個檔案。")
                    self.progress_bar['value'] = 100
                    if total_found == 0 and self.progress_bar['maximum'] == 100:
                        self.progress_label.config(text="進度: 100%")
                    self.enable_controls()
                elif msg_type == "search_stopped":
                    self.update_status("搜尋已停止。")
                    self.enable_controls()
                elif msg_type == "error":
                    self.update_status(f"錯誤: {data}")
                    self.enable_controls()

        except queue.Empty: pass
        finally: self.master.after(100, self.process_queue)

    def sort_results_by_matches(self):
        items = []
        for item in self.tree.get_children(""):
            try:
                match_val = self.tree.item(item, "values")[1]
                match_count = int(match_val) if match_val else 0
                items.append((item, match_count))
            except (ValueError, IndexError):
                print(f"Warning: Could not parse match count for item {item}, treating as 0 for sorting.")
                items.append((item, 0))

        items.sort(key=lambda x: x[1], reverse=True)

        for index, (item, _) in enumerate(items):
            self.tree.move(item, "", index)

    # on_tree_click is modified later (after _search_file_content for context)

    def safe_queue_put(self, item):
        try:
           self.update_queue.put(item)
        except Exception as e:
            print(f"Error putting item in queue: {e}")

    def _search_worker(self):
        folder = self.folder_path.get()
        kw1 = self.keyword1.get().strip()
        kw2 = self.keyword2.get().strip()
        keyword_logic = self.keyword_logic.get() # MODIFIED: Get keyword logic
        whole = self.whole_word.get()
        case_sens = self.case_sensitive.get()
        content_only = self.content_only.get()
        selected_filter = self.file_type_filter.get()
        include_subfolders = self.include_subfolders.get()

        found_files_count = 0
        processed_files_count = 0
        files_to_process = []

        allowed_extensions = self.supported_extensions[selected_filter]

        self.safe_queue_put({"type": "status", "data": "正在掃描檔案列表..."})
        try:
            if include_subfolders:
                for root, _, files in os.walk(folder):
                    if self.stop_search_flag.is_set(): break
                    for filename in files:
                        if self.stop_search_flag.is_set(): break
                        if filename.startswith('~$'): continue
                        _, ext = os.path.splitext(filename.lower())
                        if ext not in allowed_extensions: continue
                        file_path = os.path.join(root, filename)
                        try:
                            with open(file_path, 'rb') as f: f.read(1)
                            files_to_process.append(file_path)
                        except Exception as e: print(f"警告：掃描時無法訪問檔案 {file_path}: {e}"); self.safe_queue_put({"type": "status", "data": f"警告：無法訪問，跳過 {os.path.basename(file_path)}"})
                    if self.stop_search_flag.is_set(): break
            else:
                for filename in os.listdir(folder):
                    if self.stop_search_flag.is_set(): break
                    if filename.startswith('~$'): continue
                    file_path = os.path.join(folder, filename)
                    if not os.path.isfile(file_path): continue
                    _, ext = os.path.splitext(filename.lower())
                    if ext not in allowed_extensions: continue
                    try:
                        with open(file_path, 'rb') as f: f.read(1)
                        files_to_process.append(file_path)
                    except Exception as e: print(f"警告：無法訪問檔案 {file_path}: {e}"); self.safe_queue_put({"type": "status", "data": f"警告：無法訪問，跳過 {filename}"})

            if self.stop_search_flag.is_set(): self.safe_queue_put({"type": "search_stopped", "data": None}); return
        except Exception as e: self.safe_queue_put({"type": "error", "data": f"掃描資料夾時出錯: {e}"}); self.safe_queue_put({"type": "search_complete", "data": 0}); return

        total_files = len(files_to_process)
        self.safe_queue_put({"type": "progress", "data": (0, total_files)})
        if total_files > 0: self.safe_queue_put({"type": "status", "data": f"找到 {total_files} 個符合類型且可讀的檔案，開始讀取內容..."})
        else: self.safe_queue_put({"type": "status", "data": "未找到符合類型且可讀的檔案。"})

        for file_path in files_to_process:
            if self.stop_search_flag.is_set(): self.safe_queue_put({"type": "search_stopped", "data": None}); return
            processed_files_count += 1
            self.safe_queue_put({"type": "progress", "data": (processed_files_count, total_files)})
            if processed_files_count % 10 == 0 or processed_files_count == total_files:
                 self.safe_queue_put({"type": "status", "data": f"正在處理: {os.path.basename(file_path)} ({processed_files_count}/{total_files})"})

            try:
                # MODIFIED: Pass keyword_logic
                match_found, location_info, kw1_count, kw2_count = self._search_file_content(
                    file_path, kw1, kw2, case_sens, whole, content_only, keyword_logic
                )
                if match_found:
                    found_files_count += 1
                    _, ext = os.path.splitext(file_path.lower())
                    file_type = "Excel" if ext in self.supported_extensions['Excel'] else "PDF"
                    self.safe_queue_put({"type": "add_result", "data": (file_path, location_info, kw1_count, kw2_count, file_type)})
            except Exception as e:
                print(f"錯誤：處理檔案 '{file_path}' 時發生未預期錯誤: {e}")
                self.safe_queue_put({"type": "status", "data": f"錯誤：處理檔案 '{os.path.basename(file_path)}' 時出錯: {type(e).__name__}"})

        self.safe_queue_put({"type": "search_complete", "data": found_files_count})

    def _check_single_keyword(self, text_chunk, keyword, case_sensitive, whole_word):
        count = 0
        if not keyword or not text_chunk: return False, 0
        text_str = str(text_chunk)
        if not text_str.strip(): return False, 0
        try:
            if whole_word:
                flags = 0 if case_sensitive else re.IGNORECASE
                pattern = r'\b' + re.escape(keyword) + r'\b'
                matches = re.finditer(pattern, text_str, flags=flags)
                count = sum(1 for _ in matches)
                return count > 0, count
            else:
                if not case_sensitive:
                    keyword_lower = keyword.lower()
                    text_lower = text_str.lower()
                    start = 0
                    while True:
                        idx = text_lower.find(keyword_lower, start)
                        if idx == -1: break
                        count += 1
                        start = idx + len(keyword_lower)
                else:
                    start = 0
                    while True:
                        idx = text_str.find(keyword, start)
                        if idx == -1: break
                        count += 1
                        start = idx + len(keyword)
                return count > 0, count
        except re.error as re_err:
            print(f"Warning: Regex error (Keyword: '{keyword}'): {re_err}")
            if not case_sensitive: count = text_str.lower().count(keyword.lower())
            else: count = text_str.count(keyword)
            return count > 0, count
        except Exception as e:
            print(f"Warning: Error during keyword check (Keyword: '{keyword}', Text type: {type(text_chunk)}): {e}")
            return False, 0

    # MODIFIED: _search_file_content signature and logic
    def _search_file_content(self, file_path, kw1, kw2, case_sens, whole, content_only, keyword_logic):
        _, ext = os.path.splitext(file_path.lower())
        first_kw1_location = None
        first_kw2_location = None # MODIFIED: Added for kw2 location
        kw1_found_in_file = False
        kw2_found_in_file = False # MODIFIED: Initialize to False
        kw1_total_count = 0
        kw2_total_count = 0
        workbook = None
        pdf_document = None

        kw1_strip = kw1.strip() # Use stripped versions for checks
        kw2_strip = kw2.strip()

        if not kw1_strip: # Should be caught by UI, but as a safeguard
            return False, "錯誤: 關鍵字1不能為空", 0, 0

        try:
            if ext == '.xlsx' and HAS_OPENPYXL:
                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=content_only)
                except Exception as load_err:
                    print(f"Error: Failed to load XLSX file: {file_path}, Error: {load_err}")
                    return False, f"XLSX 讀取錯誤: {type(load_err).__name__}", 0, 0

                for sheet_name in workbook.sheetnames:
                    if self.stop_search_flag.is_set(): break
                    try:
                        sheet = workbook[sheet_name]
                        if sheet.sheet_state != 'visible': continue
                        if sheet.max_row == 0 or sheet.max_column == 0: continue

                        for row_idx, row in enumerate(sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column), 1):
                            if self.stop_search_flag.is_set(): break
                            for col_idx, cell in enumerate(row, 1):
                                if cell.value is None: continue
                                cell_text = str(cell.value)
                                if not cell_text.strip(): continue

                                kw1_match_in_cell, count1 = self._check_single_keyword(cell_text, kw1_strip, case_sens, whole)
                                if kw1_match_in_cell:
                                    kw1_found_in_file = True
                                    kw1_total_count += count1
                                    if first_kw1_location is None:
                                        first_kw1_location = f"Excel Sheet '{sheet_name}', Cell {get_excel_column_letter(col_idx)}{row_idx} (關鍵字1)"

                                if kw2_strip: # Only search for kw2 if it's provided
                                    kw2_match_in_cell, count2 = self._check_single_keyword(cell_text, kw2_strip, case_sens, whole)
                                    if kw2_match_in_cell:
                                        kw2_found_in_file = True
                                        kw2_total_count += count2
                                        if first_kw2_location is None:
                                            first_kw2_location = f"Excel Sheet '{sheet_name}', Cell {get_excel_column_letter(col_idx)}{row_idx} (關鍵字2)"
                            if self.stop_search_flag.is_set(): break
                    except Exception as sheet_err:
                        print(f"Warning: Error reading XLSX Sheet '{sheet_name}' in file '{file_path}': {sheet_err}")
                        continue
                if workbook: workbook.close(); workbook = None

            elif ext == '.xls' and HAS_XLRD:
                try:
                    workbook = xlrd.open_workbook(file_path, on_demand=True, logfile=open(os.devnull, 'w'))
                except Exception as load_err:
                     print(f"Error: Failed to load XLS file: {file_path}, Error: {load_err}")
                     return False, f"XLS 讀取錯誤: {type(load_err).__name__}", 0, 0

                for sheet_idx in range(workbook.nsheets):
                    if self.stop_search_flag.is_set(): break
                    sheet_name = "Unknown"
                    try:
                        sheet = workbook.sheet_by_index(sheet_idx)
                        sheet_name = sheet.name
                        if sheet.visibility != 0: continue
                        if sheet.nrows == 0 or sheet.ncols == 0: continue

                        for row_idx in range(sheet.nrows):
                            if self.stop_search_flag.is_set(): break
                            for col_idx in range(sheet.ncols):
                                try:
                                    cell_value = sheet.cell_value(row_idx, col_idx)
                                    if cell_value is None: continue
                                    cell_text = str(cell_value)
                                    if not cell_text.strip(): continue

                                    kw1_match_in_cell, count1 = self._check_single_keyword(cell_text, kw1_strip, case_sens, whole)
                                    if kw1_match_in_cell:
                                        kw1_found_in_file = True
                                        kw1_total_count += count1
                                        if first_kw1_location is None:
                                            col_letter = '?'
                                            try: col_letter = get_excel_column_letter(col_idx + 1)
                                            except NameError:
                                                if HAS_XLRD_FORMULA:
                                                    try: col_letter = xlrd.formula.cellname(row_idx, col_idx).split('$')[-1].rstrip('0123456789')
                                                    except: pass
                                            first_kw1_location = f"Excel Sheet '{sheet_name}', Cell {col_letter}{row_idx+1} (關鍵字1)"

                                    if kw2_strip: # Only search for kw2 if it's provided
                                        kw2_match_in_cell, count2 = self._check_single_keyword(cell_text, kw2_strip, case_sens, whole)
                                        if kw2_match_in_cell:
                                            kw2_found_in_file = True
                                            kw2_total_count += count2
                                            if first_kw2_location is None:
                                                col_letter_kw2 = '?'
                                                try: col_letter_kw2 = get_excel_column_letter(col_idx + 1)
                                                except NameError:
                                                    if HAS_XLRD_FORMULA:
                                                        try: col_letter_kw2 = xlrd.formula.cellname(row_idx, col_idx).split('$')[-1].rstrip('0123456789')
                                                        except: pass
                                                first_kw2_location = f"Excel Sheet '{sheet_name}', Cell {col_letter_kw2}{row_idx+1} (關鍵字2)"
                                except Exception as cell_err:
                                     print(f"Warning: Error reading XLS cell ({row_idx},{col_idx}) in sheet '{sheet_name}', file '{file_path}': {cell_err}")
                                     continue
                            if self.stop_search_flag.is_set(): break
                    except Exception as sheet_err:
                        print(f"Warning: Error reading XLS Sheet '{sheet_name}' (index {sheet_idx}) in file '{file_path}': {sheet_err}")
                        continue

            elif ext == '.pdf' and HAS_PYMUPDF:
                try:
                    pdf_document = fitz.open(file_path)
                    if pdf_document.is_encrypted:
                        print(f"Warning: PDF file is password protected, cannot read: {file_path}")
                        try: pdf_document.close()
                        except: pass
                        return False, "PDF 受密碼保護", 0, 0
                    if pdf_document.page_count == 0:
                        print(f"Warning: PDF file has no pages: {file_path}")
                        try: pdf_document.close()
                        except: pass
                        return False, "PDF 沒有頁面", 0, 0

                    for page_num in range(pdf_document.page_count):
                        if self.stop_search_flag.is_set(): break
                        page = pdf_document[page_num]
                        try:
                            page_text = page.get_text("text")
                            if not page_text or not page_text.strip(): continue

                            kw1_match_in_page, count1 = self._check_single_keyword(page_text, kw1_strip, case_sens, whole)
                            if kw1_match_in_page:
                                kw1_found_in_file = True
                                kw1_total_count += count1
                                if first_kw1_location is None: first_kw1_location = f"PDF 頁面 {page_num + 1} (關鍵字1)"

                            if kw2_strip: # Only search for kw2 if it's provided
                                kw2_match_in_page, count2 = self._check_single_keyword(page_text, kw2_strip, case_sens, whole)
                                if kw2_match_in_page:
                                    kw2_found_in_file = True
                                    kw2_total_count += count2
                                    if first_kw2_location is None: first_kw2_location = f"PDF 頁面 {page_num + 1} (關鍵字2)"
                        except Exception as page_err:
                            print(f"Warning: Error reading PDF page {page_num + 1} in file '{file_path}': {page_err}")
                            continue
                    if pdf_document: pdf_document.close(); pdf_document = None
                except Exception as pdf_err:
                    print(f"Error: Failed to open or process PDF file: {file_path}, Error: {pdf_err}")
                    if pdf_document:
                        try: pdf_document.close(); pdf_document = None
                        except: pass
                    return False, f"PDF 讀取錯誤: {type(pdf_err).__name__}", 0, 0

            # --- MODIFIED: Final check for match based on keyword_logic ---
            final_match_result = False
            if not kw2_strip: # Only keyword1 is active
                final_match_result = kw1_found_in_file
            else: # Both keywords potentially active
                if keyword_logic == "AND":
                    final_match_result = kw1_found_in_file and kw2_found_in_file
                elif keyword_logic == "OR":
                    final_match_result = kw1_found_in_file or kw2_found_in_file
                else: # Should not happen, default to AND
                    final_match_result = kw1_found_in_file and kw2_found_in_file

            current_location_info = "未找到匹配"
            if final_match_result:
                if kw1_found_in_file and first_kw1_location:
                    current_location_info = first_kw1_location
                elif kw2_found_in_file and first_kw2_location: # kw1 not found (or no location) but kw2 found
                    current_location_info = first_kw2_location
                else: # Match found but no specific cell location captured, provide generic
                    if kw1_found_in_file : current_location_info = "位置未知 (關鍵字1)"
                    elif kw2_found_in_file : current_location_info = "位置未知 (關鍵字2)"
                    else: current_location_info = "位置未知 (匹配)" # Fallback
            else:
                # If the file does not meet the final match criteria, ensure counts are 0 for this file's result.
                kw1_total_count = 0
                kw2_total_count = 0
                # current_location_info remains "未找到匹配"

            return final_match_result, current_location_info, kw1_total_count, kw2_total_count

        except Exception as e:
            print(f"Error: Unexpected error processing file '{file_path}': {e.__class__.__name__}: {e}")
            try:
                if workbook and hasattr(workbook, 'close'): workbook.close()
                if pdf_document and hasattr(pdf_document, 'close'): pdf_document.close()
            except Exception as close_err:
                print(f"Warning: Error during resource cleanup for file '{file_path}': {close_err}")
            return False, f"處理錯誤: {type(e).__name__}", 0, 0
        finally:
             try:
                if workbook and hasattr(workbook, 'close'): workbook.close()
                if pdf_document and hasattr(pdf_document, 'close'): pdf_document.close()
             except Exception as close_err:
                 print(f"Warning: Error during final resource cleanup for file '{file_path}': {close_err}")


    def on_tree_click(self, event): # MODIFIED: Popup message for location
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return
        item_id = self.tree.identify_row(event.y)
        if not item_id: return

        if item_id in self.result_details:
            filepath, location_info, kw1_count, kw2_count = self.result_details[item_id]
            if not os.path.exists(filepath): messagebox.showerror("檔案錯誤", f"檔案不存在或已被移動/刪除:\n{filepath}"); return
            try:
                if platform.system() == "Windows": os.startfile(filepath)
                elif platform.system() == "Darwin": os.system(f'open "{filepath}"')
                else: os.system(f'xdg-open "{filepath}"')

                popup_title = "檔案詳情"
                popup_message = f"檔案: {os.path.basename(filepath)}\n"
                # MODIFIED: More generic location info display
                if location_info and location_info not in ["未找到匹配", "位置未知 (匹配)", "位置未知 (關鍵字1)", "位置未知 (關鍵字2)"]:
                     popup_message += f"首個匹配位置 (近似): {location_info}\n"
                popup_message += f"關鍵字1 ('{self.keyword1.get()}') 出現次數: {kw1_count}\n"
                kw2_value = self.keyword2.get().strip()
                if kw2_value:
                    popup_message += f"關鍵字2 ('{kw2_value}') 出現次數: {kw2_count}\n"

                messagebox.showinfo(popup_title, popup_message, parent=self.master)

            except Exception as e: messagebox.showerror("開啟檔案錯誤", f"無法開啟檔案:\n{filepath}\n\n錯誤: {e}")

# --- Main execution block ---
if __name__ == "__main__":
    missing_libs = []
    if not HAS_OPENPYXL: missing_libs.append("openpyxl (.xlsx)")
    if not HAS_XLRD: missing_libs.append("xlrd (.xls)")
    if not HAS_PYMUPDF: missing_libs.append("PyMuPDF (.pdf)")

    root = tk.Tk()
    root.withdraw()

    if missing_libs:
         warn_msg = "警告：缺少以下必要的函式庫，對應的檔案類型將無法搜尋：\n\n"
         warn_msg += "\n".join(missing_libs)
         warn_msg += "\n\n請在命令提示字元(cmd)或終端機(terminal)執行 'pip install <函式庫名稱>' 來安裝它們。\n範例: pip install openpyxl\n\n程式仍可啟動，但功能受限。"
         messagebox.showwarning("缺少函式庫", warn_msg, parent=None)

    root.deiconify()
    app = FileSearchApp(root)
    root.mainloop()
